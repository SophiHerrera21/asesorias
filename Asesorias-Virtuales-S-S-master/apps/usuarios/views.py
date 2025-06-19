from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Q, Count, Sum, Avg, Min, Max
from django.core.mail import send_mail
from django.conf import settings
from django.utils import timezone
from .models import Usuario, CodigoRecuperacion, Asesor, Aprendiz, Coordinador, Grupo, Prueba, Notificacion, Reunion, Configuracion
from apps.pqrs.models import PQRS
from .forms import UsuarioForm, SolicitarRecuperacionForm, CodigoRecuperacionForm, NuevaContrasenaForm, PerfilAprendizForm, PasswordChangeForm, AsesorRegistroForm, AsesorPerfilForm
from django import forms
from django.contrib.auth import authenticate, login, logout
from django.http import HttpResponseForbidden, JsonResponse, HttpResponse, FileResponse
from django.contrib.auth import update_session_auth_hash
from apps.pqrs.models import RespuestaPQRS
from apps.pruebas.models import EntregaPrueba
from django.utils.http import urlsafe_base64_decode
from django.contrib.auth.hashers import make_password
import json
from django.db import models
from apps.componentes.models import Componente
from asesorias_virtuales.utils.email_utils import (
    send_welcome_email,
    send_group_assignment_email,
    send_pqrs_notification,
    send_meeting_link_email,
    send_group_max_reached_notification,
    send_test_notification,
    send_account_blocked_email,
    send_account_unlock_request,
    send_error_report
)
from datetime import datetime, timedelta
from openpyxl import Workbook
from django.db.models.functions import Cast
from io import BytesIO
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment
import os
import matplotlib.pyplot as plt
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Table, TableStyle, Paragraph, Image, Spacer
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.fonts import addMapping
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from django.contrib.auth import logout as auth_logout
from django.urls import reverse
from django.views.decorators.http import require_POST
from apps.grupos.models import Grupo  # Usar el modelo correcto

def es_coordinador(user):
    return user.role == 'coordinador' and hasattr(user, 'coordinador')

def es_aprendiz(user):
    return user.role == 'aprendiz' and hasattr(user, 'aprendiz')

def es_asesor(user):
    return user.role == 'asesor' and hasattr(user, 'asesor')

@login_required
@user_passes_test(es_coordinador)
def lista_usuarios(request):
    # Obtener parámetros de búsqueda
    query = request.GET.get('q', '')
    tipo = request.GET.get('tipo', '')
    estado = request.GET.get('estado', '')
    
    # Filtrar usuarios
    usuarios = Usuario.objects.all()
    
    if query:
        usuarios = usuarios.filter(
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query) |
            Q(documento__icontains=query)
        )
    
    if tipo:
        usuarios = usuarios.filter(role=tipo)
    
    if estado:
        usuarios = usuarios.filter(is_active=(estado == 'activo'))
    
    context = {
        'usuarios': usuarios,
        'query': query,
        'tipo': tipo,
        'estado': estado,
        'tipos_usuario': Usuario.ROLES,
    }
    return render(request, 'usuarios/lista_usuarios.html', context)

@login_required
@user_passes_test(es_coordinador)
def crear_usuario(request):
    if request.method == 'POST':
        form = UsuarioForm(request.POST)
        if form.is_valid():
            # Crear usuario sin guardar primero
            usuario = form.save(commit=False)
            usuario.username = usuario.email
            # Establecer la contraseña que el usuario ingresó
            usuario.set_password(form.cleaned_data['password1'])
            usuario.save()
            
            # Enviar correo de bienvenida usando el nuevo sistema
            from asesorias_virtuales.utils.email_utils import send_welcome_email
            send_welcome_email(usuario)
            
            messages.success(request, 'Usuario creado exitosamente. Se ha enviado un correo de bienvenida.')
            return redirect('lista_usuarios')
    else:
        form = UsuarioForm()
    
    return render(request, 'usuarios/form_usuario.html', {'form': form, 'accion': 'Crear'})

@login_required
@user_passes_test(es_coordinador)
def editar_usuario(request, usuario_id):
    usuario = get_object_or_404(Usuario, id=usuario_id)
    
    if request.method == 'POST':
        form = UsuarioForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(request, 'Usuario actualizado exitosamente.')
            return redirect('lista_usuarios')
    else:
        form = UsuarioForm(instance=usuario)
    
    return render(request, 'usuarios/form_usuario.html', {
        'form': form,
        'accion': 'Editar',
        'usuario': usuario
    })

@login_required
@user_passes_test(es_coordinador)
def eliminar_usuario(request, usuario_id):
    usuario = get_object_or_404(Usuario, id=usuario_id)
    
    try:
        # Enviar correo de notificación
        send_mail(
            subject='S&S Asesorías - Cuenta eliminada',
            message=f'''
            Hola {usuario.get_full_name()},
            
            Tu cuenta ha sido eliminada del sistema.
            
            Saludos,
            Equipo S&S
            ''',
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[usuario.email],
            fail_silently=False,
        )
        
        usuario.delete()
        messages.success(request, 'Usuario eliminado exitosamente.')
    except Exception as e:
        messages.error(request, f'Error al eliminar el usuario: {str(e)}')
    
    return redirect('lista_usuarios')

@login_required
@user_passes_test(es_coordinador)
def toggle_estado_usuario(request, usuario_id):
    usuario = get_object_or_404(Usuario, id=usuario_id)
    
    try:
        usuario.is_active = not usuario.is_active
        usuario.save()
        
        # Enviar correo de notificación
        estado = 'activada' if usuario.is_active else 'desactivada'
        if not usuario.is_active:
            send_account_blocked_email(usuario)
        send_mail(
            subject=f'S&S Asesorías - Cuenta {estado}',
            message=f'''
            Hola {usuario.get_full_name()},
            
            Tu cuenta ha sido {estado}.
            
            Saludos,
            Equipo S&S
            ''',
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[usuario.email],
            fail_silently=False,
        )
        
        messages.success(request, f'Usuario {"activado" if usuario.is_active else "desactivado"} exitosamente.')
    except Exception as e:
        messages.error(request, f'Error al cambiar el estado del usuario: {str(e)}')
    
    return redirect('lista_usuarios')

def home(request):
    if request.user.is_authenticated:
        try:
            # Verificar que el usuario tenga un rol válido
            if not request.user.role:
                messages.error(request, 'Tu cuenta no tiene un rol asignado. Por favor, contacta al administrador.')
                return render(request, 'usuarios/home.html')
            # Redirigir según el rol solo si no estamos ya en una página de dashboard
            current_path = request.path
            if request.user.role == 'aprendiz' and not current_path.startswith('/usuarios/dashboard_aprendiz'):
                return redirect('usuarios:dashboard_aprendiz')
            elif request.user.role == 'asesor' and not current_path.startswith('/usuarios/dashboard_asesor'):
                return redirect('usuarios:dashboard_asesor')
            elif request.user.role == 'coordinador' and not current_path.startswith('/usuarios/dashboard_coordinador'):
                return redirect('usuarios:dashboard_coordinador')
            else:
                # Si el usuario no tiene rol válido, cerrar sesión
                logout(request)
                messages.error(request, 'Tu cuenta no tiene un rol válido. Por favor, contacta al administrador.')
                return render(request, 'usuarios/home.html')
        except Exception as e:
            # En caso de error, mostrar la página home en lugar de redirigir
            messages.error(request, f'Error al procesar tu solicitud: {str(e)}')
            return render(request, 'usuarios/home.html')
    # Si el usuario no está autenticado, mostrar home
    return render(request, 'usuarios/home.html')

# Vista para solicitar recuperación
def solicitar_recuperacion(request):
    if request.method == 'POST':
        form = SolicitarRecuperacionForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            try:
                usuario = Usuario.objects.get(email=email)
                # Generar código y guardar
                codigo = CodigoRecuperacion.generar_codigo()
                CodigoRecuperacion.objects.create(usuario=usuario, codigo=codigo)
                # Enviar correo
                send_mail(
                    subject='Recuperación de contraseña S&S',
                    message=f"Tu código de recuperación de contraseña es {codigo}",
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[usuario.email],
                    fail_silently=False,
                )
                request.session['recupera_usuario_id'] = usuario.id
                messages.success(request, 'Se ha enviado un código de recuperación a tu correo.')
                return redirect('usuarios:ingresar_codigo')
            except Usuario.DoesNotExist:
                form.add_error('email', 'No existe una cuenta con ese correo.')
    else:
        form = SolicitarRecuperacionForm()
    return render(request, 'usuarios/recuperar_solicitar.html', {'form': form})

# Vista para ingresar código
def ingresar_codigo(request):
    usuario_id = request.session.get('recupera_usuario_id')
    if not usuario_id:
        return redirect('usuarios:solicitar_recuperacion')
    usuario = get_object_or_404(Usuario, id=usuario_id)
    if request.method == 'POST':
        form = CodigoRecuperacionForm(request.POST)
        if form.is_valid():
            codigo = form.cleaned_data['codigo']
            try:
                cod = CodigoRecuperacion.objects.filter(usuario=usuario, codigo=codigo, usado=False).latest('creado')
                if cod.es_valido():
                    request.session['codigo_ok'] = True
                    messages.success(request, 'Código válido. Ahora puedes crear una nueva contraseña.')
                    return redirect('usuarios:nueva_contrasena')
                else:
                    form.add_error('codigo', 'El código es inválido o ha expirado. Solicita uno nuevo.')
            except CodigoRecuperacion.DoesNotExist:
                form.add_error('codigo', 'El código es incorrecto.')
    else:
        form = CodigoRecuperacionForm()
    return render(request, 'usuarios/recuperar_codigo.html', {'form': form, 'usuario': usuario})

# Vista para registrar nueva contraseña
def nueva_contrasena(request):
    usuario_id = request.session.get('recupera_usuario_id')
    if not usuario_id or not request.session.get('codigo_ok'):
        return redirect('usuarios:solicitar_recuperacion')
    usuario = get_object_or_404(Usuario, id=usuario_id)
    if request.method == 'POST':
        form = NuevaContrasenaForm(request.POST)
        if form.is_valid():
            nueva = form.cleaned_data['nueva_contrasena']
            usuario.set_password(nueva)
            usuario.save()
            # Marcar código como usado
            CodigoRecuperacion.objects.filter(usuario=usuario, usado=False).update(usado=True)
            # Limpiar sesión
            del request.session['recupera_usuario_id']
            del request.session['codigo_ok']
            messages.success(request, 'Contraseña actualizada correctamente. Ahora puedes iniciar sesión.')
            return redirect('usuarios:login')
    else:
        form = NuevaContrasenaForm()
    return render(request, 'usuarios/recuperar_nueva.html', {'form': form, 'usuario': usuario})

def seleccionar_rol(request):
    """
    Vista para seleccionar el rol antes de registrarse.
    Muestra tarjetas para aprendiz, asesor y coordinador.
    """
    return render(request, 'usuarios/seleccionar_rol.html')

def registro_aprendiz(request):
    if request.method == 'POST':
        form = UsuarioForm(request.POST)
        if form.is_valid() and form.cleaned_data.get('role') == 'aprendiz':
            try:
                usuario = form.save(commit=False)
                usuario.role = 'aprendiz'
                usuario.username = usuario.email
                usuario.set_password(form.cleaned_data['password1'])
                usuario.save()
                from .models import Aprendiz
                ficha = request.POST.get('ficha', 'SIN_FICHA')
                programa = request.POST.get('programa', 'SIN_PROGRAMA')
                trimestre = request.POST.get('trimestre', '1')
                Aprendiz.objects.create(
                    usuario=usuario,
                    ficha=ficha,
                    programa=programa,
                    trimestre=trimestre
                )
                # Enviar correo de bienvenida
                correo_ok = send_welcome_email(usuario)
                if not correo_ok:
                    send_error_report(
                        subject='Fallo en envío de correo de bienvenida',
                        message=f'No se pudo enviar el correo de bienvenida a {usuario.email}',
                        extra_data={'usuario': usuario.email}
                    )
                    messages.warning(request, 'Registro exitoso, pero no se pudo enviar el correo de bienvenida. El administrador ha sido notificado.')
                else:
                    messages.success(request, '🎉 ¡Registro exitoso! Te hemos enviado un correo de bienvenida con información importante. Ahora puedes iniciar sesión y comenzar a disfrutar de todos los beneficios de S&S Asesorías Virtuales.')
                return redirect('usuarios:login')
            except Exception as e:
                send_error_report('Error en registro de aprendiz', str(e), request.POST.dict())
                messages.error(request, f'Error al crear el usuario: {str(e)}')
        else:
            # Mostrar errores específicos del formulario y reportar
            error_msg = '\n'.join([f'{field}: {errors}' for field, errors in form.errors.items()])
            send_error_report('Error de validación en registro de aprendiz', error_msg, request.POST.dict())
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'Error en {field}: {error}')
    else:
        form = UsuarioForm(initial={'role': 'aprendiz'})
    form.fields['role'].widget = forms.HiddenInput()
    return render(request, 'usuarios/registro_aprendiz.html', {'form': form})

def registro_asesor(request):
    if request.method == 'POST':
        form = AsesorRegistroForm(request.POST)
        if form.is_valid() and form.cleaned_data.get('role') == 'asesor':
            try:
                usuario = form.save(commit=False)
                usuario.role = 'asesor'
                usuario.username = usuario.email
                usuario.set_password(form.cleaned_data['password1'])
                usuario.save()
                
                from .models import Asesor
                Asesor.objects.create(
                    usuario=usuario,
                    especialidad=form.cleaned_data['especialidad'],
                    experiencia='',  # Se completará en el perfil
                    titulo='',       # Se completará en el perfil
                    max_grupos=4,
                    disponibilidad='', # Se completará en el perfil
                    trimestre='1'    # Valor por defecto
                )
                
                correo_ok = send_welcome_email(usuario)
                if not correo_ok:
                    send_error_report(
                        subject='Fallo en envío de correo de bienvenida',
                        message=f'No se pudo enviar el correo de bienvenida a {usuario.email}',
                        extra_data={'usuario': usuario.email}
                    )
                    messages.warning(request, 'Registro exitoso, pero no se pudo enviar el correo de bienvenida. El administrador ha sido notificado.')
                else:
                    messages.success(request, '🎉 ¡Registro exitoso! Te hemos enviado un correo de bienvenida con información importante. Ahora puedes iniciar sesión y completar tu perfil con más detalles.')
                return redirect('usuarios:login')
            except Exception as e:
                send_error_report('Error en registro de asesor', str(e), request.POST.dict())
                messages.error(request, f'Error al crear el usuario: {str(e)}')
        else:
            error_msg = '\n'.join([f'{field}: {errors}' for field, errors in form.errors.items()])
            send_error_report('Error de validación en registro de asesor', error_msg, request.POST.dict())
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'Error en {field}: {error}')
    else:
        form = AsesorRegistroForm(initial={'role': 'asesor'})
    form.fields['role'].widget = forms.HiddenInput()
    return render(request, 'usuarios/registro_asesor.html', {'form': form})

def registro_coordinador(request):
    if request.method == 'POST':
        form = UsuarioForm(request.POST)
        if form.is_valid() and form.cleaned_data.get('role') == 'coordinador':
            try:
                usuario = form.save(commit=False)
                usuario.role = 'coordinador'
                usuario.username = usuario.email
                usuario.set_password(form.cleaned_data['password1'])
                usuario.save()
                from .models import Coordinador
                Coordinador.objects.create(
                    usuario=usuario,
                    cargo=request.POST.get('cargo', 'Coordinador'),
                    departamento=request.POST.get('departamento', 'General')
                )
                correo_ok = send_welcome_email(usuario)
                if not correo_ok:
                    send_error_report(
                        subject='Fallo en envío de correo de bienvenida',
                        message=f'No se pudo enviar el correo de bienvenida a {usuario.email}',
                        extra_data={'usuario': usuario.email}
                    )
                    messages.warning(request, 'Registro exitoso, pero no se pudo enviar el correo de bienvenida. El administrador ha sido notificado.')
                else:
                    messages.success(request, '🎉 ¡Registro exitoso! Te hemos enviado un correo de bienvenida con información importante. Ahora puedes iniciar sesión y comenzar a disfrutar de todos los beneficios de S&S Asesorías Virtuales.')
                return redirect('usuarios:login')
            except Exception as e:
                send_error_report('Error en registro de coordinador', str(e), request.POST.dict())
                messages.error(request, f'Error al crear el usuario: {str(e)}')
        else:
            error_msg = '\n'.join([f'{field}: {errors}' for field, errors in form.errors.items()])
            send_error_report('Error de validación en registro de coordinador', error_msg, request.POST.dict())
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'Error en {field}: {error}')
    else:
        form = UsuarioForm(initial={'role': 'coordinador'})
    form.fields['role'].widget = forms.HiddenInput()
    return render(request, 'usuarios/registro_coordinador.html', {'form': form})

def login_view(request):
    desbloqueo_enviado = False
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        remember = request.POST.get('remember')
        
        if not email or not password:
            messages.error(request, 'Por favor, ingresa tu correo electrónico y contraseña.')
            return render(request, 'usuarios/login.html', {'desbloqueo_enviado': desbloqueo_enviado})
        
        try:
            user = Usuario.objects.get(email=email)
            
            # Verificar si la cuenta está bloqueada
            if not user.is_active:
                if 'solicitar_desbloqueo' in request.POST:
                    from asesorias_virtuales.utils.email_utils import send_account_unlock_request
                    send_account_unlock_request(user)
                    desbloqueo_enviado = True
                    messages.info(request, 'Tu cuenta está bloqueada. Se ha enviado una solicitud de desbloqueo al coordinador.')
                else:
                    messages.error(request, 'Tu cuenta está bloqueada. Puedes solicitar el desbloqueo.')
                return render(request, 'usuarios/login.html', {'desbloqueo_enviado': desbloqueo_enviado, 'email': email})
            
            # Autenticar usando email como username
            user = authenticate(request, username=email, password=password)
            
            if user is not None:
                login(request, user)
                
                # Configurar sesión según "recordar"
                if remember:
                    request.session.set_expiry(30 * 24 * 60 * 60)  # 30 días
                else:
                    request.session.set_expiry(0)  # Sesión de navegador
                
                # Actualizar último acceso
                user.ultimo_acceso = timezone.now()
                user.save()
                print (user.role)
               
                if user.role == 'aprendiz':
                    return redirect('usuarios:dashboard_aprendiz')
                elif user.role == 'asesor':
                    return redirect('usuarios:dashboard_asesor')
                elif user.role == 'coordinador':
                    return redirect('usuarios:dashboard_coordinador')
                else:
                    messages.error(request, 'Rol de usuario no válido.')
                    return redirect('usuarios:home')
            else:
                messages.error(request, 'La contraseña ingresada es incorrecta.')
                
        except Usuario.DoesNotExist:
            messages.error(request, f'No existe una cuenta con el correo {email}.')
        except Exception as e:
            messages.error(request, f'Error al iniciar sesión: {str(e)}')
            
    return render(request, 'usuarios/login.html', {'desbloqueo_enviado': desbloqueo_enviado})

@login_required
@user_passes_test(es_aprendiz)
def dashboard_aprendiz(request):
    try:
        # Verificar que el usuario tenga el objeto Aprendiz relacionado
        if not hasattr(request.user, 'aprendiz'):
            # Crear el objeto Aprendiz si no existe
            from .models import Aprendiz
            Aprendiz.objects.create(
                usuario=request.user,
                ficha='SIN_FICHA',
                programa='SIN_PROGRAMA',
                trimestre='1'
            )
            messages.info(request, 'Se ha completado tu perfil de aprendiz.')
        
        return render(request, 'usuarios/dashboard_aprendiz.html')
    except Exception as e:
        messages.error(request, f'Error al cargar el dashboard: {str(e)}')
        # En lugar de redirigir a home, mostrar el dashboard con error
        return render(request, 'usuarios/dashboard_aprendiz.html')

@login_required
@user_passes_test(es_asesor)
def dashboard_asesor(request):
    try:
        if not hasattr(request.user, 'asesor'):
            from .models import Asesor
            Asesor.objects.create(
                usuario=request.user,
                especialidad='',
                experiencia='',
                titulo='',
                max_grupos=4,
                disponibilidad='',
                trimestre='1'
            )
            messages.info(request, 'Se ha completado tu perfil de asesor.')
        asesor_usuario = request.user
        grupos = Grupo.objects.filter(asesor=asesor_usuario, estado='activo')
        total_aprendices = sum(grupo.aprendices.count() for grupo in grupos)
        pruebas_creadas = Prueba.objects.filter(asesor=asesor_usuario).count()
        pruebas_pendientes = Prueba.objects.filter(asesor=asesor_usuario, estado='pendiente').count()
        pqrs_pendientes = PQRS.objects.filter(asesor=asesor_usuario, estado='pendiente').count()
        notificaciones_no_leidas = Notificacion.objects.filter(usuario=request.user, leida=False).count()
        ahora = timezone.now()
        mañana = ahora + timedelta(days=1)
        reuniones_proximas = Reunion.objects.filter(
            grupo__in=grupos,
            fecha__gte=ahora.date(),
            fecha__lte=mañana.date(),
            estado='programada'
        ).order_by('fecha', 'hora')
        grupos_capacidad = grupos.filter(
            estado='activo'
        ).annotate(
            num_aprendices=Count('aprendices')
        ).filter(
            num_aprendices__gte=models.F('max_aprendices') - 2
        ).distinct()
        context = {
            'grupos': grupos,
            'total_aprendices': total_aprendices,
            'pruebas_creadas': pruebas_creadas,
            'pruebas_pendientes': pruebas_pendientes,
            'pqrs_pendientes': pqrs_pendientes,
            'notificaciones_no_leidas': notificaciones_no_leidas,
            'reuniones_proximas': reuniones_proximas,
            'grupos_capacidad': grupos_capacidad
        }
        return render(request, 'usuarios/dashboard_asesor.html', context)
    except Exception as e:
        messages.error(request, f'Error al cargar el dashboard: {str(e)}')
        return render(request, 'usuarios/dashboard_asesor.html', {})

@login_required
@user_passes_test(es_coordinador)
def dashboard_coordinador(request):
    try:
        # Verificar que el usuario tenga el objeto Coordinador relacionado
        if not hasattr(request.user, 'coordinador'):
            # Crear el objeto Coordinador si no existe
            from .models import Coordinador
            Coordinador.objects.create(
                usuario=request.user,
                cargo='Coordinador',
                departamento='General'
            )
            messages.info(request, 'Se ha completado tu perfil de coordinador.')
        
        # Estadísticas generales
        total_asesores = Asesor.objects.filter(activo=True).count()
        total_aprendices = Aprendiz.objects.filter(usuario__estado='activo').count()
        total_grupos = Grupo.objects.filter(estado='activo').count()
        
        # PQRS pendientes
        pqrs_pendientes = PQRS.objects.filter(estado='pendiente').order_by('-fecha_creacion')
        
        # Grupos cerca de su capacidad máxima
        grupos_capacidad = Grupo.objects.filter(
            estado='activo'
        ).annotate(
            num_aprendices=Count('aprendices')
        ).filter(
            num_aprendices__gte=models.F('max_aprendices') - 2
        ).distinct()
        
        # Asesores con más grupos
        asesores_ocupados = Asesor.objects.filter(
            activo=True,
            grupos__estado='activo'
        ).annotate(
            num_grupos=Count('grupos', filter=models.Q(grupos__estado='activo'))
        ).filter(num_grupos__gte=models.F('max_grupos') - 1)
        
        # Reuniones próximas
        ahora = timezone.now()
        mañana = ahora + timedelta(days=1)
        reuniones_proximas = Reunion.objects.filter(
            fecha__gte=ahora.date(),
            fecha__lte=mañana.date(),
            estado='programada'
        ).order_by('fecha', 'hora')
        
        context = {
            'total_asesores': total_asesores,
            'total_aprendices': total_aprendices,
            'total_grupos': total_grupos,
            'pqrs_pendientes': pqrs_pendientes,
            'grupos_capacidad': grupos_capacidad,
            'asesores_ocupados': asesores_ocupados,
            'reuniones_proximas': reuniones_proximas
        }
        return render(request, 'usuarios/dashboard_coordinador.html', context)
    except Exception as e:
        messages.error(request, f'Error al cargar el dashboard: {str(e)}')
        # En lugar de redirigir a home, mostrar el dashboard con error
        return render(request, 'usuarios/dashboard_coordinador.html', {})

@login_required
@user_passes_test(es_coordinador)
def configuracion_coordinador(request):
    if request.method == 'POST':
        # Procesar formulario de configuración institucional
        if 'nombre_institucion' in request.POST:
            config = Configuracion.objects.first()
            if not config:
                config = Configuracion()
            
            config.nombre_institucion = request.POST.get('nombre_institucion')
            config.direccion = request.POST.get('direccion')
            config.telefono = request.POST.get('telefono')
            config.email = request.POST.get('email')
            
            if 'logo' in request.FILES:
                config.logo = request.FILES['logo']
            
            config.save()
            messages.success(request, 'Configuración institucional actualizada exitosamente.')
            return redirect('usuarios:configuracion_coordinador')
        
        # Procesar formulario de parámetros generales
        elif 'tiempo_sesion' in request.POST:
            config = Configuracion.objects.first()
            if not config:
                config = Configuracion()
            
            config.tiempo_sesion = request.POST.get('tiempo_sesion')
            config.max_intentos_login = request.POST.get('max_intentos_login')
            config.tiempo_bloqueo = request.POST.get('tiempo_bloqueo')
            config.max_tamano_archivo = request.POST.get('max_tamano_archivo')
            config.tipos_archivo = request.POST.get('tipos_archivo')
            
            config.save()
            messages.success(request, 'Parámetros generales actualizados exitosamente.')
            return redirect('usuarios:configuracion_coordinador')
    
    config = Configuracion.objects.first()
    componentes = Componente.objects.all()
    
    context = {
        'config': config,
        'componentes': componentes
    }
    return render(request, 'usuarios/configuracion.html', context)

@login_required
@user_passes_test(es_coordinador)
def gestionar_componente(request, componente_id=None):
    if request.method == 'POST':
        if componente_id:
            componente = get_object_or_404(Componente, id=componente_id)
            accion = request.POST.get('accion')
            
            if accion == 'editar':
                componente.nombre = request.POST.get('nombre')
                componente.descripcion = request.POST.get('descripcion')
                componente.estado = request.POST.get('estado') == 'true'
                componente.save()
                return JsonResponse({'status': 'success', 'message': 'Componente actualizado exitosamente.'})
            
            elif accion == 'eliminar':
                componente.delete()
                return JsonResponse({'status': 'success', 'message': 'Componente eliminado exitosamente.'})
        else:
            # Crear nuevo componente
            nombre = request.POST.get('nombre')
            descripcion = request.POST.get('descripcion')
            estado = request.POST.get('estado') == 'true'
            
            componente = Componente.objects.create(
                nombre=nombre,
                descripcion=descripcion,
                estado=estado
            )
            return JsonResponse({'status': 'success', 'message': 'Componente creado exitosamente.'})
    
    elif request.method == 'GET' and componente_id:
        componente = get_object_or_404(Componente, id=componente_id)
        return JsonResponse({
            'id': componente.id,
            'nombre': componente.nombre,
            'descripcion': componente.descripcion,
            'estado': componente.estado
        })
    
    return JsonResponse({'status': 'error', 'message': 'Método no permitido.'}, status=405)

@login_required
def entrada_unica_coordinador(request):
    # Solo permite acceso a coordinadores cuyo correo contiene '+coord'
    if not (request.user.role == 'coordinador' and '+coord' in request.user.email):
        return HttpResponseForbidden('Acceso denegado: solo para coordinadores autorizados.')
    return render(request, 'usuarios/entrada_unica_coordinador.html')

@login_required
@user_passes_test(es_aprendiz)
def perfil_aprendiz(request):
    usuario = request.user
    if request.method == 'POST':
        form = PerfilAprendizForm(request.POST, request.FILES, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(request, 'Perfil actualizado correctamente')
            return redirect('perfil_aprendiz')
    else:
        form = PerfilAprendizForm(instance=usuario)
    
    # Obtener estadísticas del aprendiz
    grupos = request.user.aprendiz.grupos.all()
    pruebas_entregadas = EntregaPrueba.objects.filter(aprendiz=request.user.aprendiz).count()
    total_pruebas = EntregaPrueba.objects.filter(grupo__in=grupos).count()
    pruebas_calificadas = EntregaPrueba.objects.filter(aprendiz=request.user.aprendiz, calificacion__isnull=False).count()
    pqrs_pendientes = PQRS.objects.filter(aprendiz=request.user.aprendiz, estado='pendiente').count()
    total_pqrs = PQRS.objects.filter(aprendiz=request.user.aprendiz).count()
    
    context = {
        'form': form,
        'grupos': grupos,
        'pruebas_entregadas': pruebas_entregadas,
        'total_pruebas': total_pruebas,
        'pruebas_calificadas': pruebas_calificadas,
        'pqrs_pendientes': pqrs_pendientes,
        'total_pqrs': total_pqrs,
    }
    return render(request, 'usuarios/perfil_aprendiz.html', context)

@login_required
@user_passes_test(es_aprendiz)
def cambiar_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Contraseña actualizada exitosamente.')
            return redirect('perfil_aprendiz')
    else:
        form = PasswordChangeForm(request.user)
    
    return render(request, 'usuarios/cambiar_password.html', {'form': form})

def logout_view(request):
    logout(request)
    messages.success(request, 'Has cerrado sesión exitosamente.')
    return redirect('usuarios:home')

def password_reset_request(request):
    if request.method == 'POST':
        form = SolicitarRecuperacionForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            try:
                usuario = Usuario.objects.get(email=email)
                # Aquí deberías generar un código/token y enviarlo por correo
                # Por simplicidad, solo enviamos un mensaje
                send_mail(
                    'Recuperación de contraseña',
                    'Haz solicitado recuperar tu contraseña. Si no fuiste tú, ignora este mensaje.',
                    settings.DEFAULT_FROM_EMAIL,
                    [email],
                    fail_silently=False,
                )
                messages.success(request, 'Se ha enviado un correo de recuperación si el email existe en el sistema.')
                return redirect('usuarios:login')
            except Usuario.DoesNotExist:
                messages.success(request, 'Se ha enviado un correo de recuperación si el email existe en el sistema.')
                return redirect('usuarios:login')
    else:
        form = SolicitarRecuperacionForm()
    return render(request, 'usuarios/password_reset_request.html', {'form': form})

def password_reset_confirm(request, uidb64, token):
    try:
        uid = urlsafe_base64_decode(uidb64).decode()
        usuario = Usuario.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, Usuario.DoesNotExist):
        usuario = None

    if usuario is not None:
        # Validar el token (en este caso, el código de recuperación)
        try:
            codigo_obj = usuario.codigos_recuperacion.filter(codigo=token, usado=False).latest('creado')
        except CodigoRecuperacion.DoesNotExist:
            messages.error(request, 'El código de recuperación es inválido o ha expirado.')
            return redirect('usuarios:password_reset_done')

        if not codigo_obj.es_valido():
            messages.error(request, 'El código de recuperación ha expirado.')
            return redirect('usuarios:password_reset_done')

        if request.method == 'POST':
            form = NuevaContrasenaForm(request.POST)
            if form.is_valid():
                nueva = form.cleaned_data['nueva_contrasena']
                usuario.password = make_password(nueva)
                usuario.save()
                codigo_obj.usado = True
                codigo_obj.save()
                messages.success(request, 'Contraseña restablecida correctamente.')
                return redirect('usuarios:password_reset_complete')
        else:
            form = NuevaContrasenaForm()
        return render(request, 'usuarios/password_reset_confirm.html', {'form': form, 'usuario': usuario})
    else:
        messages.error(request, 'El enlace de recuperación no es válido.')
        return redirect('usuarios:password_reset_done')

def password_reset_done(request):
    return render(request, 'usuarios/password_reset_done.html')

def password_reset_complete(request):
    return render(request, 'usuarios/password_reset_complete.html')

@login_required
@user_passes_test(es_asesor)
def perfil_asesor(request):
    usuario = request.user
    asesor = request.user.asesor
    
    if request.method == 'POST':
        # Manejar datos del usuario
        usuario_form = PerfilAprendizForm(request.POST, request.FILES, instance=usuario)
        # Manejar datos del asesor
        asesor_form = AsesorPerfilForm(request.POST, instance=asesor)
        
        if usuario_form.is_valid() and asesor_form.is_valid():
            usuario_form.save()
            asesor_form.save()
            messages.success(request, 'Perfil actualizado correctamente.')
            return redirect('usuarios:perfil_asesor')
        else:
            # Mostrar errores si hay alguno
            for field, errors in usuario_form.errors.items():
                for error in errors:
                    messages.error(request, f'Error en {field}: {error}')
            for field, errors in asesor_form.errors.items():
                for error in errors:
                    messages.error(request, f'Error en {field}: {error}')
    else:
        usuario_form = PerfilAprendizForm(instance=usuario)
        asesor_form = AsesorPerfilForm(instance=asesor)
    
    return render(request, 'usuarios/perfil_asesor.html', {
        'usuario_form': usuario_form, 
        'asesor_form': asesor_form, 
        'usuario': usuario
    })

@login_required
@user_passes_test(es_coordinador)
def perfil_coordinador(request):
    usuario = request.user
    if request.method == 'POST':
        form = PerfilAprendizForm(request.POST, request.FILES, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(request, 'Perfil actualizado correctamente.')
            return redirect('usuarios:perfil_coordinador')
    else:
        form = PerfilAprendizForm(instance=usuario)
    return render(request, 'usuarios/perfil_coordinador.html', {'form': form, 'usuario': usuario})

@login_required
@user_passes_test(es_coordinador)
def responder_pqrs(request, pqrs_id):
    try:
        pqrs = PQRS.objects.get(id=pqrs_id)
    except PQRS.DoesNotExist:
        return redirect('usuarios:pqrs_coordinador')
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            estado = data.get('estado')
            mensaje = data.get('mensaje')
            notificar = data.get('notificar', True)
            
            if not mensaje:
                return JsonResponse({'success': False, 'error': 'El mensaje es requerido'})
            
            # Crear la respuesta
            respuesta = RespuestaPQRS.objects.create(
                pqrs=pqrs,
                usuario=request.user,
                mensaje=mensaje,
                es_coordinador=True
            )
            
            # Actualizar estado de la PQRS
            pqrs.estado = estado
            pqrs.save()
            
            # Enviar notificación por correo si está habilitado
            if notificar:
                try:
                    send_mail(
                        f'Respuesta a tu PQRS #{pqrs.id}',
                        f'Hola {pqrs.usuario.get_full_name()},\n\n'
                        f'Has recibido una respuesta a tu PQRS #{pqrs.id}:\n\n'
                        f'{mensaje}\n\n'
                        f'Estado actual: {pqrs.get_estado_display()}\n\n'
                        f'Saludos,\nEquipo S&S Asesorías Virtuales',
                        settings.DEFAULT_FROM_EMAIL,
                        [pqrs.usuario.email],
                        fail_silently=False,
                    )
                except Exception as e:
                    print(f"Error al enviar correo: {e}")
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    respuestas = RespuestaPQRS.objects.filter(pqrs=pqrs).order_by('fecha_creacion')
    return render(request, 'usuarios/responder_pqrs.html', {
        'pqrs': pqrs,
        'respuestas': respuestas
    })

@login_required
@user_passes_test(es_coordinador)
def cambiar_estado_pqrs(request, pqrs_id):
    if request.method == 'POST':
        try:
            pqrs = PQRS.objects.get(id=pqrs_id)
            data = json.loads(request.body)
            nuevo_estado = data.get('estado')
            
            if nuevo_estado not in ['pendiente', 'en_proceso', 'resuelto']:
                return JsonResponse({'success': False, 'error': 'Estado no válido'})
            
            pqrs.estado = nuevo_estado
            pqrs.save()
            
            return JsonResponse({'success': True})
        except PQRS.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'PQRS no encontrada'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

@login_required
@user_passes_test(es_coordinador)
def detalles_pqrs(request, pqrs_id):
    try:
        pqrs = PQRS.objects.get(id=pqrs_id)
        respuestas = RespuestaPQRS.objects.filter(pqrs=pqrs).order_by('fecha_creacion')
        
        data = {
            'id': pqrs.id,
            'usuario': pqrs.usuario.get_full_name(),
            'tipo': pqrs.get_tipo_display(),
            'estado': pqrs.get_estado_display(),
            'fecha': pqrs.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
            'mensaje': pqrs.mensaje,
            'respuestas': [{
                'usuario': r.usuario.get_full_name(),
                'fecha': r.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
                'mensaje': r.mensaje
            } for r in respuestas]
        }
        
        return JsonResponse(data)
    except PQRS.DoesNotExist:
        return JsonResponse({'error': 'PQRS no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@user_passes_test(es_coordinador)
def pqrs_coordinador(request):
    # Obtener filtros
    tipo = request.GET.get('tipo', '')
    estado = request.GET.get('estado', '')
    fecha = request.GET.get('fecha', '')
    
    # Filtrar PQRS
    pqrs_list = PQRS.objects.all()
    if tipo:
        pqrs_list = pqrs_list.filter(tipo=tipo)
    if estado:
        pqrs_list = pqrs_list.filter(estado=estado)
    if fecha:
        pqrs_list = pqrs_list.filter(fecha_creacion__date=fecha)
    
    # Estadísticas para gráficos
    pqrs_peticiones = PQRS.objects.filter(tipo='peticion').count()
    pqrs_quejas = PQRS.objects.filter(tipo='queja').count()
    pqrs_reclamos = PQRS.objects.filter(tipo='reclamo').count()
    pqrs_sugerencias = PQRS.objects.filter(tipo='sugerencia').count()
    
    pqrs_pendientes = PQRS.objects.filter(estado='pendiente').count()
    pqrs_en_proceso = PQRS.objects.filter(estado='en_proceso').count()
    pqrs_resueltas = PQRS.objects.filter(estado='resuelto').count()
    
    context = {
        'pqrs_list': pqrs_list,
        'pqrs_peticiones': pqrs_peticiones,
        'pqrs_quejas': pqrs_quejas,
        'pqrs_reclamos': pqrs_reclamos,
        'pqrs_sugerencias': pqrs_sugerencias,
        'pqrs_pendientes': pqrs_pendientes,
        'pqrs_en_proceso': pqrs_en_proceso,
        'pqrs_resueltas': pqrs_resueltas,
    }
    
    return render(request, 'usuarios/pqrs_coordinador.html', context)

@login_required
@user_passes_test(es_aprendiz)
def seleccionar_componentes(request):
    aprendiz = request.user
    componentes = Componente.objects.all()
    mensaje_error = ""
    mensaje_exito = ""
    componentes_no_asignados = []

    if request.method == 'POST':
        seleccionados = request.POST.getlist('componentes')
        if len(seleccionados) > 2:
            mensaje_error = "Solo puedes seleccionar hasta 2 componentes."
        elif len(seleccionados) == 0:
            mensaje_error = "Debes seleccionar al menos un componente."
        else:
            for comp_id in seleccionados:
                componente = Componente.objects.get(id=comp_id)
                # Buscar grupo activo con asesor disponible (no más de 4 grupos activos)
                grupo = Grupo.objects.filter(
                    componente=componente,
                    estado='activo',
                    asesor__grupos_asesor__estado='activo'
                ).annotate(
                    num_aprendices=models.Count('aprendices'),
                    num_grupos_asesor=models.Count('asesor__grupos_asesor', filter=models.Q(asesor__grupos_asesor__estado='activo'))
                ).filter(num_aprendices__lt=15, num_grupos_asesor__lte=4).first()
                if grupo:
                    grupo.aprendices.add(aprendiz)
                    grupo.save()
                else:
                    componentes_no_asignados.append(componente.nombre)
            if componentes_no_asignados:
                mensaje_error = f"No hay grupo o asesor disponible con cupo para: {', '.join(componentes_no_asignados)}. Por favor, elige otro componente o espera disponibilidad."
            elif not mensaje_error:
                mensaje_exito = "¡Componentes asignados y grupo encontrado exitosamente!"

    componentes_asignados = Grupo.objects.filter(aprendices=aprendiz).values_list('componente_id', flat=True)

    context = {
        'componentes': componentes,
        'componentes_asignados': componentes_asignados,
        'mensaje_error': mensaje_error,
        'mensaje_exito': mensaje_exito,
    }
    return render(request, 'usuarios/seleccionar_componentes.html', context)

@login_required
def crear_grupo(request):
    if not hasattr(request.user, 'asesor'):
        messages.error(request, 'No tienes permisos para crear grupos.')
        return redirect('home')
    
    asesor = request.user.asesor
    
    # Verificar límite de grupos
    if asesor.grupos.filter(estado='activo').count() >= asesor.max_grupos:
        messages.error(request, f'Has alcanzado el límite máximo de {asesor.max_grupos} grupos activos.')
        return redirect('dashboard_asesor')
    
    if request.method == 'POST':
        try:
            nombre = request.POST.get('nombre')
            materia = request.POST.get('materia')
            max_aprendices = int(request.POST.get('max_aprendices'))
            horario = request.POST.get('horario')
            descripcion = request.POST.get('descripcion', '')
            
            # Validaciones
            if max_aprendices > 50:
                messages.error(request, 'El número máximo de aprendices no puede ser mayor a 50.')
                return redirect('crear_grupo')
            
            grupo = Grupo.objects.create(
                nombre=nombre,
                materia=materia,
                max_aprendices=max_aprendices,
                horario=horario,
                descripcion=descripcion,
                asesor=asesor
            )
            
            messages.success(request, 'Grupo creado exitosamente.')
            return redirect('dashboard_asesor')
            
        except ValueError:
            messages.error(request, 'Por favor, ingresa valores válidos.')
            return redirect('crear_grupo')
        except Exception as e:
            messages.error(request, f'Error al crear el grupo: {str(e)}')
            return redirect('crear_grupo')
    
    return render(request, 'usuarios/crear_grupo.html')

@login_required
def crear_prueba(request, grupo_id):
    if not hasattr(request.user, 'asesor'):
        messages.error(request, 'No tienes permisos para crear pruebas.')
        return redirect('home')
    
    grupo = get_object_or_404(Grupo, id=grupo_id, asesor=request.user.asesor)
    
    if request.method == 'POST':
        try:
            titulo = request.POST.get('titulo')
            descripcion = request.POST.get('descripcion')
            fecha_limite = datetime.strptime(request.POST.get('fecha_limite'), '%Y-%m-%d %H:%M')
            puntaje_total = int(request.POST.get('puntaje_total', 100))
            tiempo_limite = int(request.POST.get('tiempo_limite', 0)) or None
            
            # Validaciones
            if fecha_limite < timezone.now():
                messages.error(request, 'La fecha límite no puede ser en el pasado.')
                return redirect('crear_prueba', grupo_id=grupo_id)
            
            if puntaje_total <= 0:
                messages.error(request, 'El puntaje total debe ser mayor a 0.')
                return redirect('crear_prueba', grupo_id=grupo_id)
            
            prueba = Prueba.objects.create(
                titulo=titulo,
                descripcion=descripcion,
                fecha_limite=fecha_limite,
                puntaje_total=puntaje_total,
                tiempo_limite=tiempo_limite,
                grupo=grupo,
                asesor=request.user.asesor
            )
            
            # Enviar notificación por correo
            send_test_notification(prueba)
            
            # Crear notificación para los aprendices
            for aprendiz in grupo.aprendices.all():
                Notificacion.objects.create(
                    usuario=aprendiz.usuario,
                    titulo=f'Nueva prueba: {titulo}',
                    mensaje=f'Se ha creado una nueva prueba para el grupo {grupo.nombre}.',
                    tipo='prueba'
                )
            
            messages.success(request, 'Prueba creada exitosamente.')
            return redirect('dashboard_asesor')
            
        except ValueError as e:
            messages.error(request, f'Error en el formato de fecha: {str(e)}')
            return redirect('crear_prueba', grupo_id=grupo_id)
        except Exception as e:
            messages.error(request, f'Error al crear la prueba: {str(e)}')
            return redirect('crear_prueba', grupo_id=grupo_id)
    
    return render(request, 'usuarios/crear_prueba.html', {'grupo': grupo})

@login_required
def crear_pqrs(request):
    if request.method == 'POST':
        try:
            asunto = request.POST.get('asunto')
            mensaje = request.POST.get('mensaje')
            tipo = request.POST.get('tipo')
            prioridad = int(request.POST.get('prioridad', 1))
            
            # Validaciones
            if not asunto or not mensaje or not tipo:
                messages.error(request, 'Todos los campos son obligatorios.')
                return redirect('crear_pqrs')
            
            if prioridad < 1 or prioridad > 5:
                messages.error(request, 'La prioridad debe estar entre 1 y 5.')
                return redirect('crear_pqrs')
            
            pqrs = PQRS.objects.create(
                asunto=asunto,
                mensaje=mensaje,
                tipo=tipo,
                prioridad=prioridad,
                usuario=request.user
            )
            
            # Enviar notificación por correo
            send_pqrs_notification(pqrs)
            
            # Crear notificación para los coordinadores
            coordinadores = Coordinador.objects.filter(activo=True)
            for coordinador in coordinadores:
                Notificacion.objects.create(
                    usuario=coordinador.usuario,
                    titulo=f'Nueva PQRS: {asunto}',
                    mensaje=f'Se ha recibido una nueva PQRS de tipo {pqrs.get_tipo_display()}.',
                    tipo='pqrs'
                )
            
            messages.success(request, 'PQRS enviada exitosamente.')
            return redirect('home')
            
        except Exception as e:
            messages.error(request, f'Error al enviar la PQRS: {str(e)}')
            return redirect('crear_pqrs')
    
    return render(request, 'usuarios/crear_pqrs.html')

@login_required
def unirse_grupo(request, grupo_id):
    if not hasattr(request.user, 'aprendiz'):
        messages.error(request, 'Solo los aprendices pueden unirse a grupos.')
        return redirect('home')
    
    grupo = get_object_or_404(Grupo, id=grupo_id, estado='activo')
    
    # Verificar si el aprendiz ya está en el grupo
    if grupo.aprendices.filter(id=request.user.aprendiz.id).exists():
        messages.warning(request, 'Ya perteneces a este grupo.')
        return redirect('dashboard_aprendiz')
    
    # Verificar capacidad del grupo
    if grupo.aprendices.count() >= grupo.max_aprendices:
        messages.error(request, 'El grupo ha alcanzado su capacidad máxima.')
        return redirect('dashboard_aprendiz')
    
    try:
        grupo.aprendices.add(request.user.aprendiz)
        
        # Verificar si el grupo alcanzó su capacidad máxima
        if grupo.aprendices.count() >= grupo.max_aprendices:
            send_group_max_reached_notification(grupo)
            
            # Notificar al coordinador
            coordinadores = Coordinador.objects.filter(activo=True)
            for coordinador in coordinadores:
                Notificacion.objects.create(
                    usuario=coordinador.usuario,
                    titulo=f'Grupo {grupo.nombre} al máximo',
                    mensaje=f'El grupo {grupo.nombre} ha alcanzado su capacidad máxima.',
                    tipo='grupo'
                )
        
        messages.success(request, 'Te has unido al grupo exitosamente.')
        return redirect('dashboard_aprendiz')
        
    except Exception as e:
        messages.error(request, f'Error al unirse al grupo: {str(e)}')
        return redirect('dashboard_aprendiz')

@login_required
def programar_reunion(request, grupo_id):
    if not hasattr(request.user, 'asesor'):
        messages.error(request, 'No tienes permisos para programar reuniones.')
        return redirect('home')
    
    grupo = get_object_or_404(Grupo, id=grupo_id, asesor=request.user.asesor, estado='activo')
    
    if request.method == 'POST':
        try:
            titulo = request.POST.get('titulo')
            fecha_str = request.POST.get('fecha')
            hora_str = request.POST.get('hora')
            duracion = int(request.POST.get('duracion'))
            descripcion = request.POST.get('descripcion')
            link_reunion = request.POST.get('link_reunion')
            
            # Validaciones
            if not all([titulo, fecha_str, hora_str, duracion, link_reunion]):
                messages.error(request, 'Todos los campos son obligatorios.')
                return redirect('programar_reunion', grupo_id=grupo_id)
            
            # Convertir fecha y hora
            fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
            hora = datetime.strptime(hora_str, '%H:%M').time()
            
            # Validar fecha y hora
            if fecha < timezone.now().date():
                messages.error(request, 'La fecha de la reunión no puede ser en el pasado.')
                return redirect('programar_reunion', grupo_id=grupo_id)
            
            if duracion < 15 or duracion > 180:
                messages.error(request, 'La duración debe estar entre 15 y 180 minutos.')
                return redirect('programar_reunion', grupo_id=grupo_id)
            
            # Verificar superposición de horarios
            reuniones_superpuestas = Reunion.objects.filter(
                grupo=grupo,
                fecha=fecha,
                hora__lte=hora,
                hora__gte=hora - timedelta(minutes=duracion)
            ).exists()
            
            if reuniones_superpuestas:
                messages.error(request, 'Ya existe una reunión programada en este horario.')
                return redirect('programar_reunion', grupo_id=grupo_id)
            
            reunion = Reunion.objects.create(
                titulo=titulo,
                fecha=fecha,
                hora=hora,
                duracion=duracion,
                descripcion=descripcion,
                link_reunion=link_reunion,
                grupo=grupo,
                asesor=request.user.asesor
            )
            
            # Agregar participantes
            for aprendiz in grupo.aprendices.all():
                reunion.participantes.add(aprendiz)
                
                # Crear notificación
                Notificacion.objects.create(
                    usuario=aprendiz.usuario,
                    titulo=f'Nueva reunión: {titulo}',
                    mensaje=f'Se ha programado una nueva reunión para el grupo {grupo.nombre}.',
                    tipo='reunion'
                )
            
            # Enviar notificación por correo
            send_meeting_link_email(reunion)
            
            messages.success(request, 'Reunión programada exitosamente.')
            return redirect('dashboard_asesor')
            
        except ValueError as e:
            messages.error(request, f'Error en el formato de fecha u hora: {str(e)}')
            return redirect('programar_reunion', grupo_id=grupo_id)
        except Exception as e:
            messages.error(request, f'Error al programar la reunión: {str(e)}')
            return redirect('programar_reunion', grupo_id=grupo_id)
    
    return render(request, 'usuarios/programar_reunion.html', {'grupo': grupo})

@login_required
def listar_reuniones(request, grupo_id):
    grupo = get_object_or_404(Grupo, id=grupo_id, estado='activo')
    
    # Verificar permisos
    if hasattr(request.user, 'asesor') and grupo.asesor != request.user.asesor:
        messages.error(request, 'No tienes permisos para ver estas reuniones.')
        return redirect('home')
    elif hasattr(request.user, 'aprendiz') and not grupo.aprendices.filter(id=request.user.aprendiz.id).exists():
        messages.error(request, 'No tienes permisos para ver estas reuniones.')
        return redirect('home')
    
    # Filtrar reuniones según el estado
    estado = request.GET.get('estado', '')
    if estado:
        reuniones = Reunion.objects.filter(grupo=grupo, estado=estado).order_by('fecha', 'hora')
    else:
        reuniones = Reunion.objects.filter(grupo=grupo).order_by('fecha', 'hora')
    
    context = {
        'grupo': grupo,
        'reuniones': reuniones,
        'estado_actual': estado
    }
    return render(request, 'usuarios/listar_reuniones.html', context)

@login_required
def detalle_reunion(request, reunion_id):
    reunion = get_object_or_404(Reunion, id=reunion_id)
    
    # Verificar permisos
    if hasattr(request.user, 'asesor') and reunion.asesor != request.user.asesor:
        messages.error(request, 'No tienes permisos para ver esta reunión.')
        return redirect('home')
    elif hasattr(request.user, 'aprendiz') and not reunion.participantes.filter(id=request.user.aprendiz.id).exists():
        messages.error(request, 'No tienes permisos para ver esta reunión.')
        return redirect('home')
    
    # Marcar notificación como leída si existe
    if hasattr(request.user, 'aprendiz'):
        Notificacion.objects.filter(
            usuario=request.user,
            tipo='reunion',
            titulo__startswith=f'Nueva reunión: {reunion.titulo}'
        ).update(leida=True, fecha_lectura=timezone.now())
    
    context = {
        'reunion': reunion,
        'participantes': reunion.participantes.all()
    }
    return render(request, 'usuarios/detalle_reunion.html', context)

@login_required
def gestionar_asesores(request):
    if not hasattr(request.user, 'coordinador'):
        messages.error(request, 'No tienes permisos para acceder a esta página.')
        return redirect('home')
    
    asesores = Asesor.objects.all().order_by('usuario__first_name')
    
    if request.method == 'POST':
        asesor_id = request.POST.get('asesor_id')
        accion = request.POST.get('accion')
        
        try:
            asesor = Asesor.objects.get(id=asesor_id)
            
            if accion == 'activar':
                asesor.activo = True
                messages.success(request, f'Asesor {asesor.usuario.get_full_name()} activado exitosamente.')
            elif accion == 'desactivar':
                asesor.activo = False
                messages.success(request, f'Asesor {asesor.usuario.get_full_name()} desactivado exitosamente.')
            elif accion == 'actualizar':
                asesor.especialidad = request.POST.get('especialidad')
                asesor.experiencia = request.POST.get('experiencia')
                asesor.titulo = request.POST.get('titulo')
                asesor.max_grupos = request.POST.get('max_grupos')
                asesor.disponibilidad = request.POST.get('disponibilidad')
                messages.success(request, f'Asesor {asesor.usuario.get_full_name()} actualizado exitosamente.')
            
            asesor.save()
            
        except Asesor.DoesNotExist:
            messages.error(request, 'Asesor no encontrado.')
        except Exception as e:
            messages.error(request, f'Error al procesar la solicitud: {str(e)}')
    
    return render(request, 'usuarios/gestionar_asesores.html', {'asesores': asesores})

@login_required
def gestionar_grupos(request):
    if not hasattr(request.user, 'coordinador'):
        messages.error(request, 'No tienes permisos para acceder a esta página.')
        return redirect('home')
    
    grupos = Grupo.objects.all().order_by('-fecha_creacion')
    
    if request.method == 'POST':
        grupo_id = request.POST.get('grupo_id')
        accion = request.POST.get('accion')
        
        try:
            grupo = Grupo.objects.get(id=grupo_id)
            
            if accion == 'activar':
                grupo.activo = True
                messages.success(request, f'Grupo {grupo.nombre} activado exitosamente.')
            elif accion == 'desactivar':
                grupo.activo = False
                messages.success(request, f'Grupo {grupo.nombre} desactivado exitosamente.')
            elif accion == 'reasignar':
                nuevo_asesor_id = request.POST.get('nuevo_asesor')
                nuevo_asesor = Asesor.objects.get(id=nuevo_asesor_id)
                
                # Verificar que el nuevo asesor no exceda su límite de grupos
                if nuevo_asesor.grupos.filter(estado='activo').count() >= nuevo_asesor.max_grupos:
                    messages.error(request, f'El asesor {nuevo_asesor.usuario.get_full_name()} ha alcanzado su límite de grupos.')
                    return redirect('gestionar_grupos')
                
                grupo.asesor = nuevo_asesor
                messages.success(request, f'Grupo {grupo.nombre} reasignado exitosamente.')
            
            grupo.save()
            
            # Notificar al asesor
            if accion == 'reasignar':
                Notificacion.objects.create(
                    usuario=nuevo_asesor.usuario,
                    titulo=f'Grupo reasignado: {grupo.nombre}',
                    mensaje=f'Has sido asignado como asesor del grupo {grupo.nombre}.',
                    tipo='grupo'
                )
                send_group_assignment_email(nuevo_asesor, grupo)
            
        except Grupo.DoesNotExist:
            messages.error(request, 'Grupo no encontrado.')
        except Asesor.DoesNotExist:
            messages.error(request, 'Asesor no encontrado.')
        except Exception as e:
            messages.error(request, f'Error al procesar la solicitud: {str(e)}')
    
    return render(request, 'usuarios/gestionar_grupos.html', {'grupos': grupos})

@login_required
def gestionar_pqrs(request):
    if not hasattr(request.user, 'coordinador'):
        messages.error(request, 'No tienes permisos para acceder a esta página.')
        return redirect('home')
    
    pqrs_list = PQRS.objects.all().order_by('-fecha_creacion')
    
    if request.method == 'POST':
        pqrs_id = request.POST.get('pqrs_id')
        accion = request.POST.get('accion')
        
        try:
            pqrs = PQRS.objects.get(id=pqrs_id)
            
            if accion == 'responder':
                respuesta = request.POST.get('respuesta')
                if not respuesta:
                    messages.error(request, 'La respuesta no puede estar vacía.')
                    return redirect('gestionar_pqrs')
                
                pqrs.respuesta = respuesta
                pqrs.estado = 'resuelto'
                pqrs.fecha_respuesta = timezone.now()
                pqrs.save()
                
                # Notificar al usuario
                Notificacion.objects.create(
                    usuario=pqrs.usuario,
                    titulo=f'Respuesta a tu PQRS: {pqrs.asunto}',
                    mensaje=f'Has recibido una respuesta a tu PQRS.',
                    tipo='pqrs'
                )
                
                messages.success(request, 'Respuesta enviada exitosamente.')
            
            elif accion == 'cambiar_estado':
                nuevo_estado = request.POST.get('nuevo_estado')
                pqrs.estado = nuevo_estado
                pqrs.save()
                messages.success(request, f'Estado de PQRS actualizado a {pqrs.get_estado_display()}.')
            
        except PQRS.DoesNotExist:
            messages.error(request, 'PQRS no encontrada.')
        except Exception as e:
            messages.error(request, f'Error al procesar la solicitud: {str(e)}')
    
    return render(request, 'usuarios/gestionar_pqrs.html', {'pqrs_list': pqrs_list})

@login_required
def reportes(request):
    if not hasattr(request.user, 'coordinador'):
        messages.error(request, 'No tienes permisos para acceder a esta página.')
        return redirect('home')
    
    # Filtros
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    # Consultas base
    grupos = Grupo.objects.filter(estado='activo')
    asesores = Asesor.objects.filter(activo=True)
    aprendices = Aprendiz.objects.filter(activo=True)
    
    # Aplicar filtros de fecha si existen
    if fecha_inicio and fecha_fin:
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d')
        grupos = grupos.filter(fecha_creacion__range=[fecha_inicio, fecha_fin])
    
    # Estadísticas
    stats = {
        'total_grupos': grupos.count(),
        'total_asesores': asesores.count(),
        'total_aprendices': aprendices.count(),
        'grupos_por_materia': grupos.values('materia').annotate(total=Count('id')),
        'aprendices_por_grupo': grupos.annotate(total=Count('aprendices')),
        'asesores_por_grupo': asesores.annotate(total=Count('grupos')),
        'pqrs_por_tipo': PQRS.objects.values('tipo').annotate(total=Count('id')),
        'reuniones_por_grupo': grupos.annotate(total=Count('reuniones')),
        'pruebas_por_grupo': grupos.annotate(total=Count('pruebas'))
    }
    
    return render(request, 'usuarios/reportes.html', {'stats': stats})

@login_required
def reportes_coordinador(request):
    if not hasattr(request.user, 'coordinador'):
        messages.error(request, 'No tienes permisos para acceder a esta página.')
        return redirect('usuarios:home')
    
    # Filtros
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    # Consultas base
    grupos = Grupo.objects.all()
    asesores = Asesor.objects.all()
    aprendices = Aprendiz.objects.all()
    pqrs = PQRS.objects.all()
    
    # Aplicar filtros de fecha si existen
    if fecha_inicio and fecha_fin:
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d')
        grupos = grupos.filter(fecha_creacion__range=[fecha_inicio, fecha_fin])
        pqrs = pqrs.filter(fecha_creacion__range=[fecha_inicio, fecha_fin])
    
    # Estadísticas de Grupos
    grupos_stats = {
        'total': grupos.count(),
        'activos': grupos.filter(estado='activo').count(),
        'inactivos': grupos.filter(estado='inactivo').count(),
        'por_nombre': grupos.values('nombre').annotate(
            total=Count('id'),
            activos=Count('id', filter=models.Q(estado='activo')),
            inactivos=Count('id', filter=models.Q(estado='inactivo'))
        ),
        'capacidad': {
            'cerca_maximo': grupos.filter(
                estado='activo'
            ).annotate(
                num_aprendices=Count('aprendices')
            ).filter(
                num_aprendices__gte=models.F('max_aprendices') - 2
            ).count(),
            'al_maximo': grupos.filter(
                estado='activo'
            ).annotate(
                num_aprendices=Count('aprendices')
            ).filter(
                num_aprendices=models.F('max_aprendices')
            ).count()
        }
    }
    
    # Estadísticas de Asesores
    asesores_stats = {
        'total': asesores.count(),
        'activos': asesores.filter(activo=True).count(),
        'inactivos': asesores.filter(activo=False).count(),
        'por_especialidad': asesores.values('especialidad').annotate(
            total=Count('id'),
            activos=Count('id', filter=models.Q(activo=True))
        ),
        'grupos_por_asesor': asesores.annotate(
            total_grupos=Count('grupos'),
            grupos_activos=Count('grupos', filter=models.Q(grupos__estado='activo'))
        ).values('usuario__first_name', 'usuario__last_name', 'total_grupos', 'grupos_activos')
    }
    
    # Estadísticas de Aprendices
    aprendices_stats = {
        'total': aprendices.count(),
        'activos': aprendices.filter(usuario__estado='activo').count(),
        'inactivos': aprendices.filter(usuario__estado='inactivo').count(),
        'por_programa': aprendices.values('programa').annotate(
            total=Count('id'),
            activos=Count('id', filter=models.Q(usuario__estado='activo'))
        ),
        'por_semestre': aprendices.values('trimestre').annotate(
            total=Count('id'),
            activos=Count('id', filter=models.Q(usuario__estado='activo'))
        )
    }
    
    # Estadísticas de PQRS
    pqrs_stats = {
        'total': pqrs.count(),
        'por_tipo': pqrs.values('tipo').annotate(
            total=Count('id'),
            pendientes=Count('id', filter=models.Q(estado='pendiente')),
            en_proceso=Count('id', filter=models.Q(estado='en_proceso')),
            resueltas=Count('id', filter=models.Q(estado='resuelto'))
        ),
        'por_estado': pqrs.values('estado').annotate(
            total=Count('id')
        ),
        'tiempo_respuesta': pqrs.filter(
            fecha_respuesta__isnull=False
        ).annotate(
            tiempo_respuesta=models.F('fecha_respuesta') - models.F('fecha_creacion')
        ).aggregate(
            promedio=Avg('tiempo_respuesta'),
            minimo=Min('tiempo_respuesta'),
            maximo=Max('tiempo_respuesta')
        )
    }
    
    # Estadísticas de Asesorías
    asesorias_stats = {
        'total_reuniones': Reunion.objects.count(),
        'reuniones_por_grupo': grupos.annotate(
            total_reuniones=Count('reuniones'),
            reuniones_realizadas=Count('reuniones', filter=models.Q(reuniones__estado='finalizada')),
            reuniones_canceladas=Count('reuniones', filter=models.Q(reuniones__estado='cancelada'))
        ),
        'duracion_promedio': Reunion.objects.filter(
            estado='finalizada'
        ).aggregate(
            promedio=Avg('duracion')
        ),
        'participacion': Reunion.objects.annotate(
            total_participantes=Count('participantes')
        ).aggregate(
            promedio_participacion=Avg('total_participantes')
        )
    }
    
    context = {
        'grupos_stats': grupos_stats,
        'asesores_stats': asesores_stats,
        'aprendices_stats': aprendices_stats,
        'pqrs_stats': pqrs_stats,
        'asesorias_stats': asesorias_stats,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin
    }
    
    return render(request, 'usuarios/reportes_coordinador.html', context)

@login_required
def exportar_reporte(request):
    if not hasattr(request.user, 'coordinador'):
        messages.error(request, 'No tienes permisos para acceder a esta página.')
        return redirect('home')
    
    tipo_reporte = request.GET.get('tipo')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    if not tipo_reporte:
        messages.error(request, 'Debe especificar el tipo de reporte.')
        return redirect('reportes_coordinador')
    
    try:
        wb = Workbook()
        ws = wb.active
        # --- LOGO ---
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo-ss.png')
        if os.path.exists(logo_path):
            img = XLImage(logo_path)
            img.height = 80
            img.width = 160
            ws.add_image(img, 'A1')
        # --- TÍTULO ---
        ws.merge_cells('A5:E5')
        ws['A5'] = f'Reporte de {tipo_reporte.capitalize()} - {timezone.now().strftime("%d/%m/%Y")}'
        ws['A5'].font = Font(size=16, bold=True)
        ws['A5'].alignment = Alignment(horizontal='center')
        row_offset = 7  # Para dejar espacio al logo y título
        # --- Encabezados y datos ---
        if tipo_reporte == 'grupos':
            headers = ['Grupo', 'Nombre', 'Asesor', 'Estado', 'Capacidad', 'Aprendices', 'Fecha Creación']
            for col, header in enumerate(headers, 1):
                ws.cell(row=row_offset, column=col, value=header)
            grupos = Grupo.objects.all()
            if fecha_inicio and fecha_fin:
                fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d')
                fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d')
                grupos = grupos.filter(fecha_creacion__range=[fecha_inicio, fecha_fin])
            for row, grupo in enumerate(grupos, row_offset+1):
                ws.cell(row=row, column=1, value=grupo.nombre)
                ws.cell(row=row, column=2, value=grupo.nombre)
                ws.cell(row=row, column=3, value=grupo.asesor.usuario.get_full_name())
                ws.cell(row=row, column=4, value='Activo' if grupo.activo else 'Inactivo')
                ws.cell(row=row, column=5, value=f"{grupo.aprendices.count()}/{grupo.max_aprendices}")
                ws.cell(row=row, column=6, value=grupo.aprendices.count())
                ws.cell(row=row, column=7, value=grupo.fecha_creacion.strftime('%Y-%m-%d'))
        elif tipo_reporte == 'asesores':
            headers = ['Asesor', 'Especialidad', 'Estado', 'Grupos Activos', 'Total Grupos', 'Experiencia']
            for col, header in enumerate(headers, 1):
                ws.cell(row=row_offset, column=col, value=header)
            asesores = Asesor.objects.all()
            for row, asesor in enumerate(asesores, row_offset+1):
                ws.cell(row=row, column=1, value=asesor.usuario.get_full_name())
                ws.cell(row=row, column=2, value=asesor.especialidad)
                ws.cell(row=row, column=3, value='Activo' if asesor.activo else 'Inactivo')
                ws.cell(row=row, column=4, value=asesor.grupos.filter(estado='activo').count())
                ws.cell(row=row, column=5, value=asesor.grupos.count())
                ws.cell(row=row, column=6, value=asesor.experiencia)
        elif tipo_reporte == 'pqrs':
            headers = ['ID', 'Tipo', 'Asunto', 'Usuario', 'Estado', 'Fecha Creación', 'Fecha Respuesta', 'Tiempo Respuesta']
            for col, header in enumerate(headers, 1):
                ws.cell(row=row_offset, column=col, value=header)
            pqrs_list = PQRS.objects.all()
            if fecha_inicio and fecha_fin:
                fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d')
                fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d')
                pqrs_list = pqrs_list.filter(fecha_creacion__range=[fecha_inicio, fecha_fin])
            for row, pqrs in enumerate(pqrs_list, row_offset+1):
                ws.cell(row=row, column=1, value=pqrs.id)
                ws.cell(row=row, column=2, value=pqrs.get_tipo_display())
                ws.cell(row=row, column=3, value=pqrs.asunto)
                ws.cell(row=row, column=4, value=pqrs.usuario.get_full_name())
                ws.cell(row=row, column=5, value=pqrs.get_estado_display())
                ws.cell(row=row, column=6, value=pqrs.fecha_creacion.strftime('%Y-%m-%d %H:%M'))
                ws.cell(row=row, column=7, value=pqrs.fecha_respuesta.strftime('%Y-%m-%d %H:%M') if pqrs.fecha_respuesta else '')
                if pqrs.fecha_respuesta:
                    tiempo_respuesta = pqrs.fecha_respuesta - pqrs.fecha_creacion
                    ws.cell(row=row, column=8, value=str(tiempo_respuesta))
        # --- Guardar en buffer y responder ---
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=reporte_{tipo_reporte}_{timezone.now().strftime("%Y%m%d")}.xlsx'
        return response
    except Exception as e:
        messages.error(request, f'Error al generar el reporte: {str(e)}')
        return redirect('reportes_coordinador')

@login_required
def exportar_pdf_usuarios_componentes_grupos(request):
    from io import BytesIO
    from django.http import FileResponse
    from django.conf import settings
    import os
    from apps.usuarios.models import Usuario, Grupo
    from apps.componentes.models import Componente
    from django.db.models import Count
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfbase import pdfmetrics
    from reportlab.lib.fonts import addMapping

    # Registrar Times New Roman
    pdfmetrics.registerFont(TTFont('Times-Roman', 'times.ttf'))
    pdfmetrics.registerFont(TTFont('Times-Bold', 'timesbd.ttf'))
    addMapping('Times-Roman', 0, 0, 'Times-Roman')
    addMapping('Times-Roman', 0, 1, 'Times-Bold')

    # Datos de usuarios
    usuarios = Usuario.objects.all().order_by('role', 'last_name')
    roles = ['aprendiz', 'asesor', 'coordinador']
    roles_display = {'aprendiz': 'Aprendiz', 'asesor': 'Asesor', 'coordinador': 'Coordinador'}
    roles_counts = [usuarios.filter(role=r).count() for r in roles]
    total_usuarios = sum(roles_counts)
    roles_porcentajes = [(c / total_usuarios * 100) if total_usuarios else 0 for c in roles_counts]

    # Datos de componentes más elegidos
    componentes = Componente.objects.annotate(num_pruebas=Count('pruebas')).order_by('-num_pruebas')[:10]
    comp_labels = [c.nombre for c in componentes]
    comp_counts = [c.num_pruebas for c in componentes]
    total_comp = sum(comp_counts)
    comp_porcentajes = [(c / total_comp * 100) if total_comp else 0 for c in comp_counts]

    # Datos de grupos
    grupos = Grupo.objects.all()
    grupos_activos = grupos.filter(estado='activo').count()
    grupos_inactivos = grupos.filter(estado='inactivo').count()
    total_grupos = grupos.count()

    # Gráfica de roles
    fig1, ax1 = plt.subplots(figsize=(4, 4))
    colores_roles = ['#2980b9', '#27ae60', '#8e44ad']
    wedges1, texts1, autotexts1 = ax1.pie(roles_counts, labels=[roles_display[r] for r in roles], autopct='%1.1f%%', colors=colores_roles, startangle=90, textprops={'fontsize': 10, 'fontname': 'Times New Roman'})
    plt.title('Distribución de Usuarios por Rol', fontname='Times New Roman', fontsize=14)
    plt.tight_layout()
    grafica_roles_buffer = BytesIO()
    plt.savefig(grafica_roles_buffer, format='png', bbox_inches='tight')
    plt.close(fig1)
    grafica_roles_buffer.seek(0)

    # Gráfica de componentes
    fig2, ax2 = plt.subplots(figsize=(4, 4))
    colores_comp = plt.cm.Paired(range(len(comp_labels)))
    wedges2, texts2, autotexts2 = ax2.pie(comp_counts, labels=comp_labels, autopct='%1.1f%%', colors=colores_comp, startangle=90, textprops={'fontsize': 10, 'fontname': 'Times New Roman'})
    plt.title('Componentes Más Elegidos', fontname='Times New Roman', fontsize=14)
    plt.tight_layout()
    grafica_comp_buffer = BytesIO()
    plt.savefig(grafica_comp_buffer, format='png', bbox_inches='tight')
    plt.close(fig2)
    grafica_comp_buffer.seek(0)

    # Crear PDF
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    c.setFont('Times-Roman', 20)

    # Logo
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo-ss.png')
    if os.path.exists(logo_path):
        c.drawImage(logo_path, width/2-60, height-100, width=120, height=60, preserveAspectRatio=True, mask='auto')
    c.setFont('Times-Bold', 18)
    c.drawCentredString(width/2, height-120, 'Reporte General de Usuarios, Componentes y Grupos - S&S Asesorías Virtuales')
    c.setFont('Times-Roman', 12)
    c.drawCentredString(width/2, height-140, f'Fecha de generación: {timezone.now().strftime("%d/%m/%Y %H:%M")}')

    # Tabla de usuarios
    data_usuarios = [['Nombre', 'Correo', 'Rol', 'Estado', 'Fecha de Registro']]
    for u in usuarios:
        data_usuarios.append([
            u.get_full_name(),
            u.email,
            u.get_role_display(),
            'Activo' if u.is_active else 'Inactivo',
            u.date_joined.strftime('%d/%m/%Y')
        ])
    table_usuarios = Table(data_usuarios, repeatRows=1, colWidths=[90, 110, 60, 50, 60])
    table_usuarios.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Times-Roman'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    table_usuarios.wrapOn(c, width-80, height-300)
    table_usuarios_height = 20 * (len(data_usuarios) + 1)
    table_usuarios.drawOn(c, 40, height-180-table_usuarios_height)

    # Gráfica de roles
    c.drawImage(grafica_roles_buffer, width/2-100, height-220-table_usuarios_height-180, width=200, height=180)
    c.setFont('Times-Bold', 12)
    c.drawString(40, height-220-table_usuarios_height-200, 'Explicación de la gráfica de roles:')
    c.setFont('Times-Roman', 11)
    y = height-220-table_usuarios_height-220
    for i, r in enumerate(roles):
        c.drawString(60, y, f"{roles_display[r]}: {roles_counts[i]} usuarios ({roles_porcentajes[i]:.1f}%)")
        y -= 16

    # Tabla de componentes
    data_comp = [['Componente', 'Veces Elegido', 'Porcentaje']]
    for i, cpt in enumerate(componentes):
        data_comp.append([cpt.nombre, comp_counts[i], f"{comp_porcentajes[i]:.1f}%"])
    table_comp = Table(data_comp, repeatRows=1, colWidths=[120, 60, 60])
    table_comp.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Times-Roman'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    table_comp.wrapOn(c, width-80, y-100)
    table_comp_height = 20 * (len(data_comp) + 1)
    table_comp.drawOn(c, 40, y-40-table_comp_height)

    # Gráfica de componentes
    c.drawImage(grafica_comp_buffer, width/2-100, y-80-table_comp_height-180, width=200, height=180)
    c.setFont('Times-Bold', 12)
    c.drawString(40, y-80-table_comp_height-200, 'Explicación de la gráfica de componentes:')
    c.setFont('Times-Roman', 11)
    y2 = y-80-table_comp_height-220
    for i, cpt in enumerate(componentes):
        c.drawString(60, y2, f"{cpt.nombre}: {comp_counts[i]} veces ({comp_porcentajes[i]:.1f}%)")
        y2 -= 16

    # Tabla de grupos
    data_grupos = [['Nombre', 'Nombre', 'Asesor', 'Estado', 'Capacidad', 'Aprendices']]
    for g in grupos:
        data_grupos.append([
            g.nombre,
            g.nombre,
            g.asesor.usuario.get_full_name() if hasattr(g, 'asesor') else '-',
            'Activo' if g.activo else 'Inactivo',
            f"{getattr(g, 'max_aprendices', '-')}",
            g.aprendices.count()
        ])
    table_grupos = Table(data_grupos, repeatRows=1, colWidths=[80, 60, 80, 50, 50, 50])
    table_grupos.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Times-Roman'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    table_grupos.wrapOn(c, width-80, y2-100)
    table_grupos_height = 20 * (len(data_grupos) + 1)
    table_grupos.drawOn(c, 40, y2-40-table_grupos_height)

    c.save()
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename=f'reporte_general_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf') 

@login_required
def exportar_pdf_pqrs(request):
    from io import BytesIO
    from django.http import FileResponse
    from django.conf import settings
    import os
    from apps.usuarios.models import PQRS
    from django.db.models import Count
    # Quitar: import matplotlib.pyplot as plt
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfbase import pdfmetrics
    from reportlab.lib.fonts import addMapping
    # ... resto de la función ...

@login_required
def eliminar_cuenta(request):
    if request.method == 'POST':
        user = request.user
        auth_logout(request)
        user.delete()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'redirect_url': reverse('usuarios:home')})
        return redirect('usuarios:home')
    return render(request, 'usuarios/confirmar_eliminar_cuenta.html')

@login_required
@user_passes_test(es_aprendiz)
def componentes_aprendiz(request):
    aprendiz = request.user.aprendiz
    grupos = Grupo.objects.filter(aprendices=aprendiz, estado='activo')
    componentes = Componente.objects.filter(grupos__aprendices=aprendiz, grupos__estado='activo').distinct()
    context = {
        'componentes': componentes,
        'grupos': grupos
    }
    return render(request, 'usuarios/componentes_aprendiz.html', context)

@login_required
@user_passes_test(es_aprendiz)
def grupos_aprendiz(request):
    aprendiz = request.user.aprendiz
    grupos = Grupo.objects.filter(aprendices=aprendiz, estado='activo')
    
    context = {
        'grupos': grupos
    }
    return render(request, 'usuarios/grupos_aprendiz.html', context)

@login_required
@user_passes_test(es_aprendiz)
def pruebas_aprendiz(request):
    aprendiz = request.user.aprendiz
    grupos = Grupo.objects.filter(aprendices=aprendiz, estado='activo')
    pruebas = Prueba.objects.filter(grupo__in=grupos)
    
    context = {
        'pruebas': pruebas
    }
    return render(request, 'usuarios/pruebas_aprendiz.html', context)

@login_required
@user_passes_test(es_aprendiz)
def pqrs_aprendiz(request):
    aprendiz = request.user.aprendiz
    pqrs_list = PQRS.objects.filter(usuario=aprendiz.usuario)
    
    context = {
        'pqrs_list': pqrs_list
    }
    return render(request, 'usuarios/pqrs_aprendiz.html', context)

@login_required
@user_passes_test(es_aprendiz)
def notificaciones_aprendiz(request):
    aprendiz = request.user.aprendiz
    notificaciones = Notificacion.objects.filter(usuario=aprendiz.usuario).order_by('-fecha_creacion')
    
    context = {
        'notificaciones': notificaciones
    }
    return render(request, 'usuarios/notificaciones_aprendiz.html', context)

@login_required
@user_passes_test(es_aprendiz)
def reportes_aprendiz(request):
    aprendiz = request.user.aprendiz
    grupos = Grupo.objects.filter(aprendices=aprendiz, estado='activo')
    pruebas = Prueba.objects.filter(grupo__in=grupos)
    
    # Estadísticas
    total_pruebas = pruebas.count()
    pruebas_completadas = pruebas.filter(entregas__aprendiz=aprendiz).count()
    promedio_notas = pruebas.filter(entregas__aprendiz=aprendiz).aggregate(
        promedio=Avg('entregas__nota')
    )['promedio'] or 0
    
    context = {
        'total_pruebas': total_pruebas,
        'pruebas_completadas': pruebas_completadas,
        'promedio_notas': promedio_notas,
        'grupos': grupos
    }
    return render(request, 'usuarios/reportes_aprendiz.html', context)

@login_required
@user_passes_test(es_asesor)
def componentes_asesor(request):
    asesor = request.user.asesor
    grupos = Grupo.objects.filter(asesor=asesor, estado='activo')
    componentes = Componente.objects.filter(grupos__asesor=asesor, grupos__estado='activo').distinct()
    context = {
        'componentes': componentes,
        'grupos': grupos
    }
    return render(request, 'usuarios/componentes_asesor.html', context)

@login_required
@user_passes_test(es_asesor)
def grupos_asesor(request):
    asesor = request.user.asesor
    grupos = Grupo.objects.filter(asesor=asesor).order_by('-fecha_creacion')
    
    context = {
        'grupos': grupos
    }
    return render(request, 'usuarios/grupos_asesor.html', context)

@login_required
@user_passes_test(es_asesor)
def pruebas_asesor(request):
    asesor = request.user.asesor
    grupos = Grupo.objects.filter(asesor=asesor, estado='activo')
    pruebas = Prueba.objects.filter(grupo__in=grupos).order_by('-fecha_creacion')
    
    context = {
        'pruebas': pruebas,
        'grupos': grupos
    }
    return render(request, 'usuarios/pruebas_asesor.html', context)

@login_required
@user_passes_test(es_asesor)
def pqrs_asesor(request):
    asesor = request.user.asesor
    pqrs_list = PQRS.objects.filter(asesor=asesor).order_by('-fecha_creacion')
    
    context = {
        'pqrs_list': pqrs_list
    }
    return render(request, 'usuarios/pqrs_asesor.html', context)

@login_required
@user_passes_test(es_asesor)
def notificaciones_asesor(request):
    notificaciones = Notificacion.objects.filter(usuario=request.user).order_by('-fecha_creacion')
    
    context = {
        'notificaciones': notificaciones
    }
    return render(request, 'usuarios/notificaciones_asesor.html', context)

@login_required
@user_passes_test(es_asesor)
def reportes_asesor(request):
    asesor_usuario = request.user  # Usuario
    grupos = Grupo.objects.filter(asesor=asesor_usuario, estado='activo')
    total_aprendices = sum(grupo.aprendices.count() for grupo in grupos)
    total_pruebas = Prueba.objects.filter(grupo__in=grupos).count()
    total_reuniones = Reunion.objects.filter(grupo__in=grupos).count()
    context = {
        'total_aprendices': total_aprendices,
        'total_pruebas': total_pruebas,
        'total_reuniones': total_reuniones,
        'grupos': grupos
    }
    return render(request, 'usuarios/reportes_asesor.html', context)

@login_required
@user_passes_test(es_coordinador)
def grupos_coordinador(request):
    grupos = Grupo.objects.all().order_by('-fecha_creacion')
    
    context = {
        'grupos': grupos
    }
    return render(request, 'usuarios/grupos_coordinador.html', context)

# Perfiles
@login_required
@user_passes_test(es_aprendiz)
def perfil_aprendiz(request):
    usuario = request.user
    if request.method == 'POST':
        form = PerfilAprendizForm(request.POST, request.FILES, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(request, 'Perfil actualizado correctamente.')
            return redirect('usuarios:perfil_aprendiz')
    else:
        form = PerfilAprendizForm(instance=usuario)
    return render(request, 'usuarios/perfil_aprendiz.html', {'form': form, 'usuario': usuario})

@login_required
@user_passes_test(es_coordinador)
def perfil_coordinador(request):
    usuario = request.user
    if request.method == 'POST':
        form = PerfilAprendizForm(request.POST, request.FILES, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(request, 'Perfil actualizado correctamente.')
            return redirect('usuarios:perfil_coordinador')
    else:
        form = PerfilAprendizForm(instance=usuario)
    return render(request, 'usuarios/perfil_coordinador.html', {'form': form, 'usuario': usuario})

@login_required
def cambiar_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Contraseña cambiada correctamente.')
            return redirect('usuarios:home')
    else:
        form = PasswordChangeForm(request.user)
    return render(request, 'usuarios/cambiar_password.html', {'form': form})